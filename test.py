import shutil
import os
from openpyxl import Workbook
from openpyxl import load_workbook

class case_script_auto_create():
    def __init__(self):
        pass
    
    # 获取vdbench
    @staticmethod
    def find_vdbench_parameter(step_content:str, parameter:str) -> int:
        index = step_content.find(parameter)
        num_str = ''
        for i in range(index+len(parameter)+1, index+len(parameter)+5):
            if step_content[i].isdigit():
                num_str += step_content[i]
        return int(num_str)
    
    # 获取vdbench/fio 数据块参数
    @staticmethod
    def find_vdbench_xfersize(step_content:str, tool:str) -> str:
        index1 = step_content.find('（')
        index2 = step_content.find('）')
        res = step_content[index1+1:index2]
        if tool == 'vdbench':
            res = res.replace('，', ',25,')
            return res+',25'
        return res
    
    # 获取vdbench是否需要一致性校验
    @staticmethod
    def find_vdbench_cc(case_title:str) -> bool:
        vdbench_cc = False
        if '写' in case_title:
            vdbench_cc = True
        return vdbench_cc
    
    # 获取fio的 读写模式(rw) 设置参数
    @staticmethod
    def find_fio_rw(case_title:str) -> str:
        if '顺序读写' in case_title:
            return 'rw'
        elif '随机读写' in case_title:
            return 'randrw'
        elif '顺序读' in case_title:
            return 'read'
        elif '随机读' in case_title:
            return 'randread'
        elif '顺序写' in case_title:
            return 'write'
        elif '随机写' in case_title:
            return 'randwrite'
        else:
            pass
    # ----------------------------------------------------------------------------------------------------
    
    # 主函数
    def main(self):
        wb = load_workbook('D:\\Sugon_Work\openpyxl_script_create\\基础IO_0815_612.xlsx', read_only=True)
        #print(wb.sheetnames)
        ws = wb.active
        case_row_index = input("输入测试用例行号：")

        # 1. case description info get
        script_name = ws['B{}'.format(case_row_index)].value
        case_number = ws['A{}'.format(case_row_index)].value
        case_title = ws['E{}'.format(case_row_index)].value
        description = case_title
        test_category = ws['K{}'.format(case_row_index)].value
        check_point = ws['L{}'.format(case_row_index)].value
        step_raw_info = ws['O{}'.format(case_row_index)].value    # steps原始信息，需处理
        step_raw = step_raw_info.split('\n')
        
        # 1.1 抽取测试用例中 vdbench/fio parameters
        vdbench_rdpct = self.find_vdbench_parameter(step_raw_info, 'rdpct')
        fio_rwmixread = vdbench_rdpct
        print('vdbench/fio 读写比例设置为：{}'.format(vdbench_rdpct))
        
        vdbench_xfersize = self.find_vdbench_xfersize(step_raw_info, tool)
        fio_bssplit = vdbench_xfersize
        print('vdbench/fio 块大小设置为：{}'.format(vdbench_xfersize))

        vdbench_seekpct = self.find_vdbench_parameter(step_raw_info, 'seekpct')
        fio_seekpct = vdbench_seekpct
        print('vdbench 随机比例设置为：{}'.format(vdbench_seekpct))
    
        # 2. case model get
        tool = input('输入测试工具: ')
        script_path = os.getcwd()    # 获取当前路径
        if tool == 'v' or tool == 'vdbench':
            source = script_path+'\case_content_vdb_model.py'
        elif tool == 'f' or tool == 'fio':
            source = script_path+'\case_content_fio_model.py'
        target = script_path+'\case_script'
        shutil.copy(source, target)
        
        f_model = open(source, 'r', encoding='UTF-8')
        slist = f_model.readlines()

        # 3. modified data
        # 3.1 修改脚本的注释描述内容
        f = open(target, 'r', encoding='UTF-8')
        flist = f.readlines()
        # f = open(target, 'w', encoding='UTF-8')
        flist[3] = 'case number: {}\n'.format(case_number)
        flist[4] = 'case title: {}\n'.format(case_title)
        flist[5] = 'test category: {}\n'.format(test_category)
        flist[6] = 'check point: {}\n'.format(check_point)
        
        # step 内容需要特殊处理
        flist[12] = '@steps: {}\n'.format(step_raw[0])
        raw_num = 12
        temp_str = ''
        i = 1
        while i < len(step_raw):
            raw_num += 1
            if len(step_raw[i]) > 70:    # 需要加行
                temp = ''
                step_long_raw = step_raw[i].split('，')
                for x in range(len(step_long_raw)):
                    if len(temp) + len(step_long_raw[x]) < 70:
                        temp += (step_long_raw[x] + '，')
                    elif len(temp) + len(step_long_raw[x]) >= 70:
                        # flist[raw_num] = '        {}\n'.format(temp)
                        flist.insert(raw_num, '        {}\n'.format(temp))
                        temp = step_long_raw[x]+'，'
                        raw_num += 1
                flist[raw_num] = '        {}\n'.format(temp)
            else:
                flist[raw_num] = '        {}\n'.format(step_raw[i])
            i += 1
        flist[raw_num+1] = '@changelog:\n'
        flist[raw_num+2] = '"""\n'
        # 复制引用库部分
        raw_num += 4
        for i in range(17, 22):
            flist[raw_num] = slist[i]
            raw_num += 1
        raw_num += 1
        # 修改脚本类名
        script_class_name = input("输入脚本类名：")
        flist[class_name_raw_num] = 'class {}(BasicioJBODScriptBase):\n'.format(script_class_name)
        raw_num += 2
            
        #f.writelines(flist)
        #f.close()
        
        
        # 3.2 设置测试用例脚本参数
        """
        for i in range(raw_num, len(flist)):
            if 'class' in flist[i]:    # 类名设置语句
                class_name_raw_num = i
            if 'super' in flist[i]:    # 基类设置参数语句，下一句为用例参数设置首行。这里为27，比实际行号少1
                super_para_raw_num = i
                continue
        """
        # 复制不同部分（工具）
        if tool == 'v' or tool == 'vdbench':
            vdbench_cc = self.find_vdbench_cc(case_title)
            print('vdbench一致性校验：{}'.format(vdbench_cc))
            for i in range(raw_num, 50):
                if 'use' in flist[i]:
                    flist[i] = "        cls.vdbench_parameters_dict['use_vdbench'] = {}\n".format(True)
                elif 'elapsed' in flist[i]:
                    flist[i] = "        cls.vdbench_parameters_dict['elapsed'] = '{}'\n".format(30)
                elif 'rdpct' in flist[i]:
                    flist[i] = "        cls.vdbench_parameters_dict['rdpct'] = '{}'\n".format(vdbench_rdpct)
                elif 'seekpct' in flist[i]:
                    flist[i] = "        cls.vdbench_parameters_dict['seekpct'] = '{}'\n".format(vdbench_seekpct)
                elif 'xfersize' in flist[i]:
                    flist[i] = "        cls.vdbench_parameters_dict['xfersize'] = '({})'\n".format(vdbench_xfersize)
                elif 'check' in flist[i]:
                    flist[i] = "        cls.vdbench_parameters_dict['consistency_check'] = {}\n".format(vdbench_cc)
            """
            flist[super_para_raw_num+1] = "        cls.vdbench_parameters_dict['use_vdbench'] = {}\n".format(True)
            flist[super_para_raw_num+2] = "        cls.vdbench_parameters_dict['elapsed'] = '{}'\n".format(30)
            flist[super_para_raw_num+3] = "        cls.vdbench_parameters_dict['seekpct'] = '{}'\n".format(vdbench_seekpct)
            flist[super_para_raw_num+4] = "        cls.vdbench_parameters_dict['rdpct'] = '{}'\n".format(vdbench_rdpct)
            flist[super_para_raw_num+5] = "        cls.vdbench_parameters_dict['xfersize'] = '({})'\n".format(vdbench_xfersize)
            flist[super_para_raw_num+6] = "        cls.vdbench_parameters_dict['consistency_check'] = {}\n".format(vdbench_cc)
            """
        elif tool == 'f' or tool == 'fio':
            for i in range(raw_num, 50):
                if 'USE' in flist[i]:
                    flist[i] = "        cls.fio_parameters_dict[FioEnum.FIO_USE.value] = {}\n".format(True)
                elif 'RW' in flist[i]:
                    flist[i] = "        cls.fio_parameters_dict[FioEnum.FIO_RW.value] = '{}'\n".format(fio_rw)
                elif 'BSSPLIT' in flist[i]:
                    flist[i] = "        cls.fio_parameters_dict[FioEnum.FIO_BSSPLIT.value] = '({})'\n".format(fio_bssplit)
                elif 'SEEKPCT' in flist[i]:
                    flist[i] = "        cls.fio_parameters_dict[FioEnum.FIO_SEEKPCT.value] = {}\n".format(fio_seekpct)
                elif 'RWMIXREAD' in flist[i]:
                    flist[i] = "        cls.fio_parameters_dict[FioEnum.FIO_RWMIXREAD.value] = '{}'\n".format(fio_rwmixread)
            """
            fio_rw = self.find_fio_rw(case_title)
            flist[super_para_raw_num+1] = "        cls.fio_parameters_dict[FioEnum.FIO_USE.value] = {}\n".format(True)
            flist[super_para_raw_num+2] = "        cls.fio_parameters_dict[FioEnum.FIO_RUNTIME.value] = '{}'\n".format(30)
            flist[super_para_raw_num+3] = "        cls.fio_parameters_dict[FioEnum.FIO_RW.value] = '{}'\n".format(fio_rw)
            flist[super_para_raw_num+4] = "        cls.fio_parameters_dict[FioEnum.FIO_BSSPLIT.value] = '({})'\n".format(fio_bssplit)
            flist[super_para_raw_num+5] = "        cls.fio_parameters_dict[FioEnum.FIO_RWMIXREAD.value] = '{}'\n".format(fio_rwmixread)
            """
        else:
            print('别闹，没这工具...')

        # 3.3 修改脚本末尾的内容
        raw_num += 
        run_raw_num = [x for x in range(45, len(flist)) if 'run' in flist[x]]
        flist[run_raw_num[0]] = '    {}.run()'.format(script_class_name)
        f = open(target, 'w', encoding='UTF-8')
        f.writelines(flist)
        f.close()
        os.rename("case_script", script_name+'.py')    # 格式化

if __name__ == "__main__":
    test = case_script_auto_create()
    test.main()


