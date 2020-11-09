import os
from openpyxl import Workbook
from openpyxl import load_workbook

pass_sum = 0
fail_sum = 0

report_dir_path = r"D:\Sugon_Work\test\esat\scripts\system_test\basic_io\parallel_io\diff_dg\jbod_jbod".split(
    '\\')
report_dir_path = "\\".join(report_dir_path)
print(report_dir_path)

with os.scandir(basepath) as entries:
    for entry in entries:
        if entry.is_dir():
            print(entry.name)

with os.scandir(report_dir_path) as entries:
    for entry in entries:
        if entry.is_dir() and 'ps3_auto_test_run' in entry.name:
            if entry.is_dir() and 'reports' in entry.name:
                if entry.is_file() and 'report' in entry.name:
                    wb = load_workbook(entry, read_only=True)    # open excel
                    # # ws = wb.get_sheet_by_name('基础IO')
                    # ws = wb['基础IO最小用例集']
                    pass_current_num = ws['F2'].value
                    fail_current_num = ws['G2'].value
                    print('该次测试的pass数目为：%d' % pass_current_num)
                    print('该次测试的fail数目为：%d' % fail_current_num)
                    pass_sum += pass_current_num
                    fail_sum += fail_current_num

print('该vm_下的所有近期所有测试pass总数为：%d' % pass_sum)
print('该vm_下的所有近期所有测试fail总数为：%d' % fail_sum)



