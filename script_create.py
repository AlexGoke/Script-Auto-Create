"""
author: liuyuan
description: 根据测试用例excel自动脚本生成工具
             主体框架，根据需要在子类中实现 testscene 针对性的具体特殊功能
version: 1.0 / 2020.08.25 / basicio-jbod-vdbench/fio 自动生成
         1.1 / 2020.09.20 / basicio-raid-vdbench/fio 自动生成
                            basicio-raid&jbod-vdbench/fio 自动生成
         1.2 / 2020.10.08 / auto-rename 自动重命名工具
                            自动匹配excel的各列内容
                            适应最新的脚本名称、格式规则

"""

import shutil
import os
import abc
import subprocess

import text_template
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from expand_function import FuncSet


class case_script_auto_create(metaclass=abc.ABCMeta):

    excel = None
    # 每次生成一类脚本前需要修改的信息 全局变量
    # 键盘输入
    case_row_index = None       # 该用例的excel行号
    tool = ''                   # 该用例使用的测试工具
    global script_class_name    # 该用例的脚本类名
    # 子类规定
    template = ''
    # test_scene_para = []              # 需要查找的测试场景信息
    need_test_tool_para_list = []       # 需要查找的测试工具信息

    flist = []                  # 该用例的脚本内容
    target = None               # 最终生成的脚本对象

    # 脚本头部注释内容
    script_name = None
    case_number = None
    case_title = None
    description = None
    test_category = None
    check_point = None
    test_scene_info = None
    step_raw_info = ''
    step_info = ''

    tool_para_dict = {}    # 该类测试用例 工具参数信息
    scene_para_dict = {}   # 该类测试用例 场景参数信息

    def __init__(self):
        pass

    @abc.abstractclassmethod
    def prepara(cls) -> None:
        """
        description:  脚本开始自动生成前完成获取相关信息准备工作，打开用例文件及输入必要的用例信息
        parametr：    None
        return：      None
        """
        pass

    @classmethod
    def prepara_base(cls) -> None:
        """
        description:  生成器的基础准备工作，所有脚本类型需准备的相同的部分
        parametr：    None
        return：      None
        """
        wb = load_workbook('./模拟平台联调测试用例_生成用_1008.xlsx', read_only=True)
        sheet = wb.get_sheet_by_name('基础IO')
        # print(wb.sheetnames)
        # sheet = wb['基础IO']
        # print(sheet.title)
        cls.excel = sheet
        # cls.excel = wb.active    # 这个是获取当前正在显示的sheet！巨坑，删掉
        # 每次生成一类脚本前需要修改的信息 全局变量
        tool_referce = ['v', 'f']
        try:
            cls.tool = input('输入测试工具 [v/f]: ')
            if not (cls.tool in tool_referce):
                raise ValueError(
                    'The tool name can only be selected from either v or f')
        except ValueError as error:
            print(repr(error))
            cls.tool = input('重新选择测试工具 [v/f]：')
        cls.script_class_name = input("输入脚本类名：")
        # cls.need_test_tool_para_list = ['rdpct', 'seekpct', 'offset', 'align', 'range', 'xfersize']

    @ classmethod
    def case_excel_access(cls, need_test_tool_para_list: list, case_row_index: int) -> None:
        """
        description: 测试用例的excel中各种信息获取
        parameter:   need_test_tool_para_list: 想要在测试用例的excel中获取到的参数名称列表
                     case_row_index: 测试用例所在excel的行号
        return：     None
        """
        ws = cls.excel
        # 查找列名说明行
        line = None
        for row in range(1, 20):
            if ws['A%d' % row].value == '测试编号':
                line = row
        # 自动识别列名
        column_name_row = ws[line]
        # print(column_name_row)
        column_dict = dict()
        for cell in column_name_row:
            # print(row.coordinate)
            if cell.value:
                # print(cell.row, get_column_letter(cell.column))
                column_dict[cell.value] = get_column_letter(cell.column)
        # print(column_dict)
        # cls.case_number = ws['A{}'.format(case_row_index)].value
        cls.case_number = ws[column_dict['测试编号'] + str(case_row_index)].value
        # cls.script_name = ws['B{}'.format(case_row_index)].value
        cls.script_name = ws[column_dict['脚本编号'] + str(case_row_index)].value
        # cls.case_title = ws['F{}'.format(case_row_index)].value
        cls.case_title = ws[column_dict['用例标题'] + str(case_row_index)].value
        cls.description = cls.case_title
        # cls.test_category = ws['K{}'.format(case_row_index)].value
        cls.test_category = ws[column_dict['分类'] + str(case_row_index)].value
        # cls.check_point = ws['L{}'.format(case_row_index)].value
        cls.check_point = ws[column_dict['检查项/测试点'] +
                             str(case_row_index)].value
        # cls.test_scene_info = ws['N{}'.format(case_row_index)].value
        cls.test_scene_info = ws[column_dict['测试场景'] +
                                 str(case_row_index)].value
        # cls.step_raw_info = ws['O{}'.format(case_row_index)].value    # steps原始信息，需处理
        cls.step_raw_info = ws[column_dict['测试步骤'] + str(case_row_index)].value
        cls.step_info = cls.step_raw_info.split('\n')

        # 抽取测试用例中 vdbench/fio common-parameters
        cls.tool_para_dict = FuncSet.find_tool_parameter(
            cls.step_raw_info, cls.need_test_tool_para_list, cls.tool, cls.case_title)
        # 抽取测试用例中 test_scene common-parameters
        cls.scene_para_dict = FuncSet.find_scene_parameter(cls.test_scene_info)

    @ classmethod
    def model_info_access(cls, template: str) -> None:
        """
        description ： 复制获取用例的模板内容
        parameter:     template 模板文件名称
        return:        None
        """
        script_path = os.getcwd()    # 获取当前路径
        source = script_path + '\\template\\' + template
        cls.target = script_path+'\case_script'
        shutil.copy(source, cls.target)    # 之后改为新建一个txt
        f_model = open(source, 'r', encoding='UTF-8')
        cls.flist = f_model.readlines()

    @ classmethod
    def script_description_content(cls) -> int:
        """
        @description : 组合完成脚本内容
        """
        # 1 修改脚本的注释描述内容
        cls.flist[3] = 'case number: {}\n'.format(cls.case_number)
        cls.flist[4] = 'case title: {}\n'.format(cls.case_title)
        cls.flist[5] = 'test category: {}\n'.format(cls.test_category)
        cls.flist[6] = 'check point: {}\n'.format(cls.check_point)

        # 2 步骤内容需要特殊处理
        # cls.flist[12] = '@steps: {}\n'.format(cls.step_info[0])
        raw_num = 14
        temp_str = ''
        i = 0
        print(cls.step_info)

        for i in range(len(cls.step_info)):
            if len(cls.step_info)-1 == i:
                cls.flist.insert(raw_num, '{}\n'.format(cls.step_info[i]))
                break
            if len(cls.step_info[i]) > 70:    # 需要加行
                # 新方法，excel中手动加入"；"分隔符
                step_long_raw = cls.step_info[i].split('；')    # 以中文分号进行拆分
                for x in range(len(step_long_raw)):
                    cls.flist.insert(raw_num, '{}\n'.format(step_long_raw[x]))
                    raw_num += 1

                # 旧方法，用算法拆
                """
                temp = ''
                # step_long_raw = cls.step_info[i].split('，')    # 测试步骤尽量用中文逗号
                # step_long_raw = cls.step_info[i].split(',')
                for x in range(len(step_long_raw)):
                    if len(temp) + len(step_long_raw[x]) < 70:
                        temp += (step_long_raw[x] + '，')
                    elif len(temp) + len(step_long_raw[x]) >= 70:
                        cls.flist.insert(raw_num, '{}\n'.format(temp))
                        temp = step_long_raw[x]+'，'
                        raw_num += 1
                cls.flist.insert(raw_num, '{}\n'.format(temp))
                """
            else:
                cls.flist.insert(
                    raw_num, '{}\n'.format(cls.step_info[i]))
                raw_num += 1
            # i += 1

        # 2 修改脚本类名
        for i in range(raw_num, len(cls.flist)):
            if 'class' in cls.flist[i]:
                run_raw_num = i
                break
        cls.flist[run_raw_num] = cls.flist[run_raw_num].replace(
            'xxx', cls.script_class_name)

    @ classmethod
    @ abc.abstractmethod
    def testscene_parameter_set(cls, flist: str, test_scene_info: str) -> None:
        """
        description: 测试场景的参数设置(抽象)——————由各类脚本生成器子类实现
        parameter： flist:           脚本内容缓存
                    test_scene_info: 测试场景信息(目前没有用到，之后会加自动解析，现在都是在测试用例脚本中写死了各种场景信息)
        """
        pass

    @ classmethod
    @ abc.abstractclassmethod
    def testtool_parameter_set(cls, flist: str, tool_para_dict: dict, tool: str) -> None:
        """
        description: 测试工具的参数设置(抽象)———————由各类脚本生成器子类实现
        parameter:  flist:           脚本内容缓存
                    tool_para_didt:  测试工具相关参数字典
                    tool:            测试工具名称
        return：None
        """
        pass

    @ classmethod
    def script_content_end(cls) -> None:
        """
        description: 修改脚本末尾的内容
        """
        text = text_template.SCRIPT_END.format(
            class_name=cls.script_class_name)
        cls.flist.append(text)

        f = open(cls.target, 'w', encoding='UTF-8')
        f.writelines(cls.flist)
        f.close()
        if '.py' in cls.script_name:    # 重命名
            final_name = cls.script_name.split('/')[-1]
            os.rename("case_script", final_name)
        else:
            final_name = cls.script_name + '.py'
            os.rename("case_script", final_name)
        cmd = "autopep8 --in-place --aggressive --aggressive {}.py".format(
            cls.script_name)
        subprocess.getoutput(cmd)
        shutil.move("./{}".format(final_name), "./product/")

# ----------------------------------------------------------------------------------------------
    @ classmethod
    def script_generate(cls):
        """
        @description  : 生成脚本——主流程
        """
        cls.case_excel_access(cls.need_test_tool_para_list, cls.case_row_index)
        print(cls.tool_para_dict)
        print(cls.template)
        cls.model_info_access(cls.template)
        cls.script_description_content()
        # 测试场景设置 ———— 根据excel种测试场景信息，获取设置相应参数
        # 先针对raid&jbod部分 添加这个函数，之后重构[各类脚本定制化]
        cls.testscene_parameter_set(cls.flist, cls.test_scene_info)
        # 测试工具设置
        cls.testtool_parameter_set(cls.flist, cls.tool_para_dict, cls.tool)
        cls.script_content_end()

    @ classmethod
    def run(cls) -> None:
        """
        description:    负责循环运行逻辑
        """
        cls.prepara()    # 打开excel，选择工具，输入类名
        temp = input("input case raw number [x] or raw number range [x-y]:")
        if '-' in temp:
            row_range = [int(x) for x in temp.split('-')]
            for i in range(row_range[0], row_range[1]+1):
                cls.case_row_index = i
                if not cls.excel['B%d' % i].value:    # 类名
                    # 重置类名，以下循环都将是该字符串
                    print('自动获取到以下的同类脚本类名为')
                    cls.script_class_name = cls.excel['A%d' % i].value
                else:
                    print(cls.script_class_name)
                    cls.script_generate()
        else:
            cls.case_row_index = temp
            cls.script_generate()


if __name__ == "__main__":
    print('目前基类已经不作为程序入口了，请创建针对脚本类型的子类脚本')
    pass
