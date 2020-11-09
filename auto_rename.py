
# 自动根据excel更改脚本文件名

import os
from openpyxl import Workbook
from openpyxl import load_workbook

wait_to_rename_list = []        # 待修改的脚本名称列表(修改名称、内容)
excel_script_name_list = []     #
excel_script_content_dict = {}    # excel中测试用例的信息（暂不用）
match_case_dict = {}            # 文件与excel


def file_name_listdir(dir_path: str) -> list:
    """
    description: 返回一个装有路径下所有'b'打头的脚本文件
    """
    file_name_list = []
    for file in os.listdir(dir_path):
        if os.path.isfile(os.path.join(dir_path, file)) and file[0] == 'b':
            file_name_list.append(file)
    return file_name_list


# 获取待rename文件
file_path = r"D:\Sugon_Work\test\esat\scripts\system_test\basic_io\jbod\jbod_x4_pt_switch".split(
    '\\')
file_path = "\\".join(file_path)
print(file_path)

wait_to_rename_list = file_name_listdir(file_path)
print("该文件夹下待修改文件名的数量为：%d" % len(wait_to_rename_list))

wb = load_workbook('./jbod_passthrough.xlsx')
# ws = wb.get_sheet_by_name('基础IO')
# ws = wb['基础IO最小用例集']
ws = wb['Sheet1']

# 获取修改目标名称
# 行号范围
excel_script_name_list = []
excel_script_num_list = []
for case_row_index in range(1, 49):
    script_name = ws['B' + str(case_row_index)].value
    script_num = ws['A' + str(case_row_index)].value
    if script_name:
        excel_script_name_list.append(script_name)
        excel_script_num_list.append(script_num)
# print(excel_script_name_list)
print("excel中相对应的文件名数量为：%d" % len(excel_script_name_list))

# ------------------------------------------------------------------------重命名脚本
# rename-匹配名称方式


def rename_match_name():
    """
    description: 将文件名与excel中名称进行匹配 重命名
    """
    match_case_dict = {}    # 名称匹配且重命名的脚本名字典
    if len(excel_script_name_list) and len(wait_to_rename_list):
        for i in range(len(excel_script_name_list)):
            for y in range(len(wait_to_rename_list)):
                origin_name_number = wait_to_rename_list[y][-7:]
                target_name_number = excel_script_name_list[i].split(
                    '/')[-1][-7:]
                print("原始文件的匹配参照标志信息:  " + origin_name_number)
                print("对应excel中的匹配参照标志信息:  " + target_name_number)
                if origin_name_number == target_name_number and os.path.exists(file_path+"\\"+wait_to_rename_list[i]):
                    print("文件名称中的序号匹配成功，重命名ing... ...")
                    os.chdir(file_path)
                    old_name = wait_to_rename_list[y]
                    new_name = excel_script_name_list[i].split('/')[-1]
                    match_case_dict[old_name] = new_name
                    os.renames(os.path.join(file_path, old_name),
                               os.path.join(file_path, new_name))
                    break
                else:
                    print("文件名称匹配失败")
    rename_report()


def rename_in_order():
    """
    description: 按目录下文件遍历的顺序 进行重命名
    """
    match_case_dict = {}    # 名称匹配且重命名的脚本名字典
    if len(excel_script_name_list) and len(wait_to_rename_list):
        for i in range(len(excel_script_name_list)):
            if os.path.exists(file_path+"\\"+wait_to_rename_list[i]):
                os.chdir(file_path)
                old_name = wait_to_rename_list[i]
                new_name = excel_script_name_list[i].split('/')[-1]
                print("文件%s重命名成功，新的文件名为%s" % (old_name, new_name))
                match_case_dict[old_name] = new_name
                os.renames(os.path.join(file_path, old_name),
                           os.path.join(file_path, new_name))
            else:
                print("%s文件不存在" % wait_to_rename_list[i])
    rename_report()

# ----------------------------------------------------------------------- 修改脚本注释里的 用例编号


# 获取修改目标 用例编号
# 行号范围
# excel_script_number_list = []
# excel_script_name_list = []
# for case_row_index in range(196, 262):
#     script_number = ws['A' + str(case_row_index)].value
#     script_name = ws['B' + str(case_row_index)].value
#     if script_name:
#         excel_script_name_list.append(script_name)
#     if script_number:
#         excel_script_number_list.append(script_number)
# # print(excel_script_name_list)
# print("excel中相对应的用例编号数量为：%d" % len(excel_script_number_list))

def amend_script_content():
    """
    description: 修改脚本文件的内容（替换某行）
    """
    match_case_dict = {}    # 名称匹配且重命名的脚本名字典
    if len(excel_script_name_list) and len(wait_to_rename_list):
        for i in range(len(excel_script_name_list)):
            for y in range(len(wait_to_rename_list)):
                origin_name_number = wait_to_rename_list[i]
                target_name_number = excel_script_name_list[y].split('/')[-1]
                # print("origin_name's index number:  " + origin_name_number)
                # print("target new name's index number:  " + target_name_number)

                target = file_path+"\\"+wait_to_rename_list[i]
                if origin_name_number == target_name_number and os.path.exists(target):
                    print("文件名称匹配成功，修改用例中的脚本编号ing... ...")
                    match_case_dict[origin_name_number] = target_name_number
                    new_str = ''
                    file_data = []
                    with open(target, "r", encoding="utf-8") as f:
                        for line in f:
                            if flag_str in line:
                                line = new_str
                            file_data.append(line)

                    with open(target, "w", encoding="utf-8") as ff:
                        for line in file_data:
                            ff.write(line)
                    break
                else:
                    print("文件名称匹配失败")
    rename_report()

# ---------------------------------------------------- jbod 文件里 加一行 passthrough = True/False


def add_script_content():
    """
    description: 在任意位置增加脚本中的内容
    """
    flag_str = 'cls.passthrough = True'
    new_str1 = '        # 需要在io过程中改变passthrough的状态\n'
    new_str2 = '        cls.passthrough_io_switch = True\n'
    # new_str3 = '        cls.passthrough = {}\n'.format(True)

    # 注意：listdir只是能遍历文件名，不是文件本身。所以还需要path去拼接出完整的文件路径
    # for file in os.listdir(file_path):
    #     file_data = []
    #     if file[0] != 'b':
    #         continue
    #     target = file_path + '\\'+file
    with os.scandir(file_path) as entries:
        for entry in entries:
            if entry.is_file() and entry.name[0] == 'b':
                i = 0
                file_data = []
                with open(entry, "r", encoding="utf-8") as f:
                    single_light = True
                    for line in f:
                        if flag_str in line and single_light:
                            # line = ''    # 当要清空某行内容时，用到
                            file_data.append(line)
                            # single_light = False
                            file_data.append(new_str1)
                            file_data.append(new_str2)
                            # if 'off' in file:
                            #     file_data.append(new_str1)
                            #     file_data.append(new_str2)
                            # elif 'on' in file:
                            #     file_data.append(new_str1)
                            #     file_data.append(new_str3)
                        else:
                            file_data.append(line)
                        i += 1
                with open(entry, "w", encoding="utf-8") as ff:
                    for line in file_data:
                        ff.write(line)


# -----------------------------------------------------------------修改 excel 中的内容
def modify_excel_content():
    """
    description: 修改excel中的某单元格的内容
    """
    for i in range(9, 25):
        print(i)
        old_content = ws['N%d' % i].value
        print(old_content)
        old_content_list = list(old_content.split('\n'))
        old_content_list.insert(3, '4. 在io测试过程中，设置控制器的直发命令开关passthrough=on')
        old_content_list[4] = '5. 进行数据一致性校验'
        old_content_list[5] = '6. 清理环境'
        print(old_content_list)
        new_content = "\n".join(old_content_list)
        print(new_content)
        ws['N%d' % i] = new_content

    for i in range(33, 49):
        print(i)
        old_content = ws['N%d' % i].value
        print(old_content)
        old_content_list = list(old_content.split('\n'))
        old_content_list.insert(3, '4. 在io测试过程中，设置控制器的直发命令开关passthrough=off')
        old_content_list[4] = '5. 进行数据一致性校验'
        old_content_list[5] = '6. 清理环境'
        print(old_content_list)
        new_content = "\n".join(old_content_list)
        print(new_content)
        ws['N%d' % i] = new_content

    wb.save('new.xlsx')


def rename_report():
    print("共匹配—修改内容的脚本个数为：{}".format(len(match_case_dict)))
    for key, value in match_case_dict.items():
        print('{key}:{value}'.format(key=key, value=value))


amend_script_content()
